import pandas as pd
import streamlit as st
import base64
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def trova_indice_intestazione(df):
    for index, row in df.iterrows():
        for value in row:
            if isinstance(value, str) and "Style Image" in value:
                return index
    raise ValueError("Intestazione non trovata.")

def estrai_dati_excel(xls, sheet_name):
    df = xls.parse(sheet_name, header=None)
    header_index = trova_indice_intestazione(df)
    df = pd.read_excel(xls, sheet_name=sheet_name, header=header_index)
    
    # Rimuovi eventuali colonne "Unnamed" subito dopo aver identificato l'intestazione
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    
    # Trova e rimuovi righe dopo "Total:" se presente
    if 'Country of Origin' in df.columns:
        total_row_index = df[df['Country of Origin'].astype(str).str.contains("Total:", na=False)].index
        if not total_row_index.empty:
            df = df.iloc[:total_row_index[0]]
    
    # Gestisci i valori mancanti nelle colonne delle taglie
    taglie_columns = [col for col in df.columns if col not in [
        "Style Image", "Style Name", "Style Number", "Color", 
        "Color Code", "Color Comment", "Style Comment", 
        "Materials", "Fabrication", "Country of Origin",
        "Sugg. Retail (EUR)", "WholeSale (EUR)", "Item Discount", 
        "Units", "Total (EUR)"
    ]]
    for col in taglie_columns:
        df[col] = df[col].fillna(0)  # Sostituisci i valori mancanti con zeri
    
    return df

def estrai_e_riordina_dati_da_tutti_sheet(file_path):
    xls = pd.ExcelFile(file_path)
    
    colonne_fisse_prima = [
        "Style Image", "Style Name", "Style Number", "Color", 
        "Color Code", "Color Comment", "Style Comment", 
        "Materials", "Fabrication", "Country of Origin"
    ]
    colonne_fisse_dopo = [
        "Sugg. Retail (EUR)", "WholeSale (EUR)", "Item Discount", 
        "Units", "Total (EUR)"
    ]
    
    colonne_taglie = set()
    
    all_extracted_data_frames = []
    for sheet_name in xls.sheet_names:
        # Ignora gli sheet che contengono nel nome "cancelled"
        if "cancelled" in sheet_name.lower():
            continue
        df = estrai_dati_excel(xls, sheet_name)
        colonne_taglie.update(set(df.columns) - set(colonne_fisse_prima) - set(colonne_fisse_dopo))
        all_extracted_data_frames.append(df.assign(Sheet=sheet_name))
    
    all_extracted_data = pd.concat(all_extracted_data_frames, ignore_index=True)
    
    ordine_completo_colonne = colonne_fisse_prima + sorted(list(colonne_taglie)) + colonne_fisse_dopo + ['Sheet']

    for col in colonne_taglie:
        all_extracted_data[col] = all_extracted_data[col].fillna(0)

    all_extracted_data = all_extracted_data[all_extracted_data["Style Image"].isna()]

    all_extracted_data = all_extracted_data.reindex(columns=ordine_completo_colonne)

    return all_extracted_data

def main():
    st.title("Caricamento e elaborazione file Excel")

    uploaded_file = st.file_uploader("Carica il file Excel", type=["xlsx"])

    if uploaded_file is not None:
        # Lettura e elaborazione dei dati
        all_extracted_data = estrai_e_riordina_dati_da_tutti_sheet(uploaded_file)

        # Rimozione delle colonne "Style Image" e formattazione delle colonne delle taglie
        all_extracted_data.drop(columns=['Style Image'], inplace=True)

        # Rimuovi le colonne delle taglie che hanno solo valori zero
        all_extracted_data = all_extracted_data.loc[:, (all_extracted_data != 0).any(axis=0)]

        # Converti il DataFrame in HTML e visualizzalo
        st.write(all_extracted_data.to_html(escape=False), unsafe_allow_html=True)

        # Download del file elaborato
        output_file_path = "extracted_data.xlsx"
        all_extracted_data.to_excel(output_file_path, index=False)
        apply_conditional_formatting(output_file_path, all_extracted_data)
        st.markdown(get_binary_file_downloader_html(output_file_path, 'Excel file'), unsafe_allow_html=True)

def apply_conditional_formatting(file_path, df):
    wb = Workbook()
    ws = wb.active

    for row in df.iterrows():
        row_index, row_data = row
        ws.append(row_data.tolist())

    yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=ws.max_column):
        for cell in row:
            if cell.value != 0:
                cell.fill = yellow_fill

    wb.save(file_path)

def get_binary_file_downloader_html(bin_file, file_label='File'):
    with open(bin_file, 'rb') as f:
        data = f.read()
    bin_str = base64.b64encode(data).decode()
    href = f'<a href="data:application/octet-stream;base64,{bin_str}" download="{os.path.basename(bin_file)}">{file_label}</a>'
    return href

if __name__ == "__main__":
    main()
